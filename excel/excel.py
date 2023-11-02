from openpyxl import load_workbook


class Excel:
    def get_excel_file_obj(self, file_name: str):
        """获得一个excel的对象"""
        return load_workbook(file_name + ".xlsx")

    def get_sheet(self, workbook, sheet_name: str):
        """获得一个sheet的对象"""
        return workbook[sheet_name]

    def get_all_sheet_names(self, workbook) -> list:
        """返回excel所有的sheet名字"""
        return workbook.sheetnames

    def write_to_cell(self, sheet, value, row: int, col: int):
        """写入单元格

        Args:
            sheet (_type_): sheet
            value (_type_): 代写入的值
            row (int): 行号。从1开始
            col (int): 列号。从1开始
        """
        sheet.cell(row, col, value)

    def del_cols_range_from_second_row(
        self, sheet, max_col: int, min_col: int = 1, min_row: int = 2
    ):
        """从第二行清空指定列的数据

        Args:
            sheet (_type_): 工作表对象
            min_col (int, optional): 开始的列
            max_col (int, optional): 结束的列
            min_row (int, optional): 最小的行
        """
        # Worksheet.iter_cols 方法会返回列
        for col in sheet.iter_cols(min_col, max_col, min_row):
            for cell in col:
                cell.value = None
        # print(f'清空了工作表{min_col}-{max_col}列的数据')

    def save_to_new_excel(self, workbook, new_excel_name: str):
        """保存成新的excel文件"""
        workbook.save("new_excel/" + new_excel_name + ".xlsx")


if __name__ == "__main__":
    e = Excel()
    f = e.get_excel_file_obj("nine")
    ws = e.get_sheet(f, "承诺情况")
